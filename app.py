from flask import Flask, request, send_file, render_template_string
import pandas as pd
import numpy as np
import openpyxl
import io
import os

app = Flask(__name__)

HTML_FORM = '''
<!DOCTYPE html>
<html lang="hi">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>Insurance Reconciliation Tool</title>
  <link href="https://fonts.googleapis.com/css2?family=Sora:wght@300;400;600;700&family=Space+Mono:wght@400;700&display=swap" rel="stylesheet"/>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

    :root {
      --bg: #0a0e1a;
      --surface: #111827;
      --card: #162032;
      --border: #1e3a5f;
      --accent: #00c2ff;
      --accent2: #0077ff;
      --accent-glow: rgba(0, 194, 255, 0.18);
      --text: #e2eaf4;
      --muted: #6b8cae;
      --success: #00e5a0;
      --error: #ff4d6d;
      --font-main: 'Sora', sans-serif;
      --font-mono: 'Space Mono', monospace;
    }

    body {
      background: var(--bg);
      font-family: var(--font-main);
      color: var(--text);
      min-height: 100vh;
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: flex-start;
      padding: 40px 16px 80px;
      position: relative;
      overflow-x: hidden;
    }

    body::before {
      content: '';
      position: fixed;
      top: -120px; left: 50%;
      transform: translateX(-50%);
      width: 700px; height: 400px;
      background: radial-gradient(ellipse at center, rgba(0,119,255,0.13) 0%, transparent 70%);
      pointer-events: none;
      z-index: 0;
    }

    .grid-bg {
      position: fixed;
      inset: 0;
      background-image:
        linear-gradient(rgba(0,194,255,0.03) 1px, transparent 1px),
        linear-gradient(90deg, rgba(0,194,255,0.03) 1px, transparent 1px);
      background-size: 40px 40px;
      z-index: 0;
      pointer-events: none;
    }

    .container {
      position: relative;
      z-index: 1;
      width: 100%;
      max-width: 680px;
    }

    /* ── HEADER ── */
    .header {
      text-align: center;
      margin-bottom: 48px;
      animation: fadeDown 0.7s ease both;
    }
    .badge {
      display: inline-block;
      background: rgba(0,194,255,0.1);
      border: 1px solid rgba(0,194,255,0.3);
      color: var(--accent);
      font-family: var(--font-mono);
      font-size: 11px;
      letter-spacing: 2px;
      text-transform: uppercase;
      padding: 5px 14px;
      border-radius: 20px;
      margin-bottom: 18px;
    }
    h1 {
      font-size: clamp(26px, 5vw, 38px);
      font-weight: 700;
      line-height: 1.15;
      background: linear-gradient(135deg, #ffffff 30%, var(--accent));
      -webkit-background-clip: text;
      -webkit-text-fill-color: transparent;
      background-clip: text;
      margin-bottom: 10px;
    }
    .subtitle {
      color: var(--muted);
      font-size: 14px;
      font-weight: 300;
      letter-spacing: 0.3px;
    }

    /* ── CARD ── */
    .card {
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 20px;
      padding: 36px 32px;
      margin-bottom: 20px;
      box-shadow: 0 8px 40px rgba(0,0,0,0.4), inset 0 1px 0 rgba(255,255,255,0.04);
      animation: fadeUp 0.6s ease both;
    }
    .card:nth-child(2) { animation-delay: 0.1s; }
    .card:nth-child(3) { animation-delay: 0.18s; }

    .card-header {
      display: flex;
      align-items: center;
      gap: 12px;
      margin-bottom: 24px;
      padding-bottom: 16px;
      border-bottom: 1px solid var(--border);
    }
    .card-icon {
      width: 40px; height: 40px;
      background: linear-gradient(135deg, var(--accent2), var(--accent));
      border-radius: 10px;
      display: flex; align-items: center; justify-content: center;
      font-size: 18px;
      flex-shrink: 0;
      box-shadow: 0 0 18px var(--accent-glow);
    }
    .card-title { font-size: 16px; font-weight: 600; color: var(--text); }
    .card-desc  { font-size: 12px; color: var(--muted); margin-top: 2px; }

    /* ── UPLOAD ZONE ── */
    .upload-zone {
      border: 2px dashed var(--border);
      border-radius: 14px;
      padding: 32px 20px;
      text-align: center;
      cursor: pointer;
      transition: all 0.25s ease;
      position: relative;
      background: rgba(255,255,255,0.01);
    }
    .upload-zone:hover, .upload-zone.dragover {
      border-color: var(--accent);
      background: var(--accent-glow);
      transform: translateY(-2px);
      box-shadow: 0 0 24px var(--accent-glow);
    }
    .upload-zone input[type="file"] {
      position: absolute; inset: 0;
      opacity: 0; cursor: pointer; width: 100%; height: 100%;
    }
    .upload-icon { font-size: 34px; margin-bottom: 10px; display: block; }
    .upload-main { font-size: 14px; font-weight: 600; color: var(--text); margin-bottom: 5px; }
    .upload-sub  { font-size: 12px; color: var(--muted); }
    .upload-formats {
      display: inline-block;
      margin-top: 10px;
      font-family: var(--font-mono);
      font-size: 10px;
      color: var(--accent);
      background: rgba(0,194,255,0.08);
      border: 1px solid rgba(0,194,255,0.2);
      padding: 3px 10px;
      border-radius: 20px;
      letter-spacing: 1px;
    }
    .file-chosen {
      margin-top: 10px;
      font-size: 12px;
      color: var(--success);
      font-family: var(--font-mono);
      display: none;
      align-items: center;
      justify-content: center;
      gap: 6px;
    }
    .file-chosen.show { display: flex; }

    /* ── INFO BOX ── */
    .info-box {
      background: rgba(0,194,255,0.05);
      border: 1px solid rgba(0,194,255,0.15);
      border-radius: 12px;
      padding: 14px 16px;
      margin-top: 20px;
      display: flex;
      gap: 10px;
      align-items: flex-start;
    }
    .info-icon { font-size: 16px; flex-shrink: 0; margin-top: 1px; }
    .info-text { font-size: 12px; color: var(--muted); line-height: 1.6; }
    .info-text strong { color: var(--accent); }

    /* ── BUTTON ── */
    .btn-wrap { margin-top: 28px; }
    button[type="submit"] {
      width: 100%;
      padding: 16px;
      background: linear-gradient(135deg, var(--accent2) 0%, var(--accent) 100%);
      color: #fff;
      border: none;
      border-radius: 12px;
      font-family: var(--font-main);
      font-size: 15px;
      font-weight: 700;
      letter-spacing: 0.5px;
      cursor: pointer;
      transition: all 0.25s ease;
      box-shadow: 0 4px 24px rgba(0,119,255,0.35);
      position: relative;
      overflow: hidden;
    }
    button[type="submit"]::before {
      content: '';
      position: absolute;
      inset: 0;
      background: linear-gradient(135deg, transparent 40%, rgba(255,255,255,0.12));
      opacity: 0;
      transition: opacity 0.25s;
    }
    button[type="submit"]:hover { transform: translateY(-2px); box-shadow: 0 8px 32px rgba(0,194,255,0.45); }
    button[type="submit"]:hover::before { opacity: 1; }
    button[type="submit"]:active { transform: translateY(0); }
    button[type="submit"]:disabled {
      opacity: 0.5; cursor: not-allowed; transform: none;
    }

    /* ── LOADING ── */
    .loading {
      display: none;
      text-align: center;
      padding: 24px;
      animation: fadeUp 0.4s ease;
    }
    .loading.show { display: block; }
    .spinner {
      width: 44px; height: 44px;
      border: 3px solid rgba(0,194,255,0.15);
      border-top-color: var(--accent);
      border-radius: 50%;
      animation: spin 0.9s linear infinite;
      margin: 0 auto 14px;
    }
    .loading-text { font-size: 13px; color: var(--muted); }

    /* ── ALERT ── */
    .alert {
      border-radius: 12px;
      padding: 14px 18px;
      font-size: 13px;
      margin-bottom: 20px;
      display: flex;
      align-items: center;
      gap: 10px;
      animation: fadeUp 0.4s ease;
    }
    .alert-error  { background: rgba(255,77,109,0.1); border: 1px solid rgba(255,77,109,0.3); color: #ff8fa3; }
    .alert-success{ background: rgba(0,229,160,0.1); border: 1px solid rgba(0,229,160,0.3); color: var(--success); }

    /* ── FOOTER ── */
    .footer {
      text-align: center;
      margin-top: 40px;
      font-size: 11px;
      color: #2a4060;
      font-family: var(--font-mono);
      letter-spacing: 1px;
      animation: fadeUp 0.8s 0.3s ease both;
    }

    /* ── ANIMATIONS ── */
    @keyframes fadeDown {
      from { opacity: 0; transform: translateY(-20px); }
      to   { opacity: 1; transform: translateY(0); }
    }
    @keyframes fadeUp {
      from { opacity: 0; transform: translateY(20px); }
      to   { opacity: 1; transform: translateY(0); }
    }
    @keyframes spin {
      to { transform: rotate(360deg); }
    }

    @media (max-width: 480px) {
      .card { padding: 24px 18px; }
    }
  </style>
</head>
<body>
<div class="grid-bg"></div>
<div class="container">

  <div class="header">
    <div class="badge">⚕ Insurance Portal</div>
    <h1>Reconciliation<br/>Processing Tool</h1>
    <p class="subtitle">Upload files · Process data · Download clean Excel output</p>
  </div>

  {% if error %}
  <div class="alert alert-error">
    <span>⚠️</span>
    <span>{{ error }}</span>
  </div>
  {% endif %}

  {% if success %}
  <div class="alert alert-success">
    <span>✅</span>
    <span>{{ success }}</span>
  </div>
  {% endif %}

  <form method="POST" enctype="multipart/form-data" id="mainForm">

    <!-- MAIN FILE -->
    <div class="card">
      <div class="card-header">
        <div class="card-icon">📋</div>
        <div>
          <div class="card-title">Main File</div>
          <div class="card-desc">Insurance reconciliation data file</div>
        </div>
      </div>

      <div class="upload-zone" id="zone1">
        <input type="file" name="main_file" id="main_file" accept=".html,.xls,.xlsx" required
               onchange="showFile(this, 'chosen1', 'zone1')"/>
        <span class="upload-icon">📁</span>
        <div class="upload-main">Click to upload or drag & drop</div>
        <div class="upload-sub">HTML ya Excel file select karein</div>
        <span class="upload-formats">.HTML &nbsp;·&nbsp; .XLS &nbsp;·&nbsp; .XLSX</span>
      </div>
      <div class="file-chosen" id="chosen1">
        <span>✅</span><span id="chosen1-name"></span>
      </div>

      <div class="info-box">
        <span class="info-icon">ℹ️</span>
        <div class="info-text">
          File mein <strong>UTR NO, VISIT_ID, SPONSOR, NET BILL AMT</strong> jaise columns hone chahiye.
          HTML aur Excel dono format supported hain.
        </div>
      </div>
    </div>

    <!-- LOOKUP FILE -->
    <div class="card">
      <div class="card-header">
        <div class="card-icon">🔍</div>
        <div>
          <div class="card-title">Lookup File</div>
          <div class="card-desc">Sponsor → Payer mapping Excel file</div>
        </div>
      </div>

      <div class="upload-zone" id="zone2">
        <input type="file" name="lookup_file" id="lookup_file" accept=".xlsx" required
               onchange="showFile(this, 'chosen2', 'zone2')"/>
        <span class="upload-icon">🗂️</span>
        <div class="upload-main">Click to upload or drag & drop</div>
        <div class="upload-sub">Payer mapping file select karein</div>
        <span class="upload-formats">.XLSX only</span>
      </div>
      <div class="file-chosen" id="chosen2">
        <span>✅</span><span id="chosen2-name"></span>
      </div>

      <div class="info-box">
        <span class="info-icon">ℹ️</span>
        <div class="info-text">
          Lookup file mein <strong>2nd, 3rd, 4th column</strong> mein
          <strong>SPONSOR, Existing, Payer</strong> data hona chahiye.
        </div>
      </div>
    </div>

    <!-- SUBMIT -->
    <div class="card" style="padding: 24px 32px;">
      <div class="btn-wrap">
        <button type="submit" id="submitBtn">
          ⚡ &nbsp; Process &amp; Download Final Excel
        </button>
      </div>

      <div class="loading" id="loadingBox">
        <div class="spinner"></div>
        <div class="loading-text">Data process ho raha hai, please wait…</div>
      </div>
    </div>

  </form>

  <div class="footer">INSURANCE RECONCILIATION TOOL &nbsp;·&nbsp; SECURE &amp; PRIVATE</div>
</div>

<script>
  function showFile(input, chosenId, zoneId) {
    const chosen = document.getElementById(chosenId);
    const nameEl = document.getElementById(chosenId + '-name');
    const zone   = document.getElementById(zoneId);
    if (input.files && input.files[0]) {
      nameEl.textContent = input.files[0].name;
      chosen.classList.add('show');
      zone.style.borderColor = 'var(--success)';
      zone.style.background  = 'rgba(0,229,160,0.05)';
    }
  }

  // Drag & drop highlight
  document.querySelectorAll('.upload-zone').forEach(zone => {
    zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('dragover'); });
    zone.addEventListener('dragleave', () => zone.classList.remove('dragover'));
    zone.addEventListener('drop', () => zone.classList.remove('dragover'));
  });

  // Form submit loader
  document.getElementById('mainForm').addEventListener('submit', function() {
    const btn = document.getElementById('submitBtn');
    const loader = document.getElementById('loadingBox');
    btn.disabled = true;
    btn.textContent = 'Processing…';
    loader.classList.add('show');
  });
</script>
</body>
</html>
'''

# =============================
# ROUTES
# =============================

@app.route("/", methods=["GET"])
def index():
    return render_template_string(HTML_FORM, error=None, success=None)


@app.route("/", methods=["POST"])
def process():
    main_file   = request.files.get("main_file")
    lookup_file = request.files.get("lookup_file")

    if not main_file or not lookup_file:
        return render_template_string(HTML_FORM,
                                      error="Dono files upload karein.",
                                      success=None)

    try:
        # =============================
        # 2️⃣ MAIN FILE READ (HTML ya Excel)
        # =============================
        main_bytes = main_file.read()
        filename   = main_file.filename.lower()

        try:
            if filename.endswith(".html") or filename.endswith(".htm"):
                tables = pd.read_html(io.BytesIO(main_bytes))
                df = tables[0]
            else:
                df = pd.read_excel(io.BytesIO(main_bytes), dtype={"UTR NO": str})
        except Exception:
            tables = pd.read_html(io.BytesIO(main_bytes))
            df = tables[0]

        # =============================
        # 3️⃣ HEADER FIX
        # =============================
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)

        # =============================
        # 4️⃣ CLEAN COLUMN NAMES
        # =============================
        df.columns = df.columns.str.strip()

        # =============================
        # 5️⃣ NUMERIC FIX
        # =============================
        numeric_cols = [
            "NET BILL AMT.", "SPONSER_AMOUNT", "CLAIM AMOUNT",
            "RECEIVED AMOUNT", "TDS AMOUNT", "WRITEOFF AMOUNT",
            "PATIENT AMOUNT", "PROCESSING FEE", "LEGITIMATE DISCOUNT",
            "DISALLOWANCE AMOUNT", "Total"
        ]
        for col in numeric_cols:
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(",", "")
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # =============================
        # 6️⃣ DATE FIX
        # =============================
        date_columns = ["FILE_SUBMISSION_DT", "UTR DATE", "INVOICE DATE", "RECONCILED DATE"]
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)

        # =============================
        # 7️⃣ FILTER UNIT_NAME
        # =============================
        if "UNIT_NAME" in df.columns:
            df = df[~df["UNIT_NAME"].isin(["Zynova", "---END---"])]

        # =============================
        # 8️⃣ UTR NO FINAL FIX
        # =============================
        if "UTR NO" in df.columns:
            df["UTR NO"] = df["UTR NO"].fillna("").astype(str)
            df["UTR NO"] = df["UTR NO"].replace("nan", "").str.strip()

        # =============================
        # 9️⃣ VISIT COLUMN
        # =============================
        if "VISIT_ID" in df.columns:
            df["VISIT"] = df["VISIT_ID"].astype(str).str[:2]
            df["VISIT"] = df["VISIT"].replace("ER", "OP")
            df["VISIT_ID"] = df["VISIT_ID"].astype(str).str.replace("^ER", "OP", regex=True)

        # =============================
        # 🔟 LOOKUP FILE
        # =============================
        lookup_bytes = lookup_file.read()
        lookup_df = pd.read_excel(io.BytesIO(lookup_bytes))
        lookup_df.columns = lookup_df.columns.str.strip()
        lookup_df = lookup_df.iloc[:, 1:4]
        lookup_df.columns = ["SPONSOR", "Existing", "Payer"]

        payer_map    = lookup_df.drop_duplicates(subset="SPONSOR").set_index("SPONSOR")["Payer"]
        existing_map = lookup_df.drop_duplicates(subset="SPONSOR").set_index("SPONSOR")["Existing"]

        if "SPONSOR" in df.columns:
            df["Payer"]    = df["SPONSOR"].map(payer_map).fillna("NA")
            df["Existing"] = df["SPONSOR"].map(existing_map).fillna("NA")

        # =============================
        # 1️⃣1️⃣ TOTAL COLUMN
        # =============================
        amount_cols = [
            "RECEIVED AMOUNT", "TDS AMOUNT", "WRITEOFF AMOUNT",
            "PATIENT AMOUNT", "PROCESSING FEE", "LEGITIMATE DISCOUNT",
            "DISALLOWANCE AMOUNT"
        ]
        for col in amount_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        df["Total"] = df[[c for c in amount_cols if c in df.columns]].sum(axis=1)

        # =============================
        # 1️⃣2️⃣ FINAL COLUMN ORDER
        # =============================
        final_cols = [
            "UNIT_NAME", "RECONCILED DATE", "VISIT", "VISIT_ID", "ADMISSION NUMBER",
            "MRNO", "PATIENT NAME", "INVOICE_NO", "INVOICE DATE", "Payer", "Existing",
            "SPONSOR", "UTR NO", "UTR DATE", "NET BILL AMT.", "SPONSER_AMOUNT",
            "CLAIM AMOUNT", "RECEIVED AMOUNT", "TDS AMOUNT", "WRITEOFF AMOUNT",
            "PATIENT AMOUNT", "PROCESSING FEE", "LEGITIMATE DISCOUNT",
            "DISALLOWANCE AMOUNT", "Total", "REMARKS", "FILE_SUBMISSION_DT",
            "IS RESUBMISION", "ADMITTING DR.", "SPECIALITY"
        ]
        final_cols = [c for c in final_cols if c in df.columns]
        final_df = df[final_cols]

        # =============================
        # 1️⃣3️⃣ TEMP SAVE → OPENPYXL FORMATTING
        # =============================
        temp_buffer = io.BytesIO()
        final_df.to_excel(temp_buffer, index=False)
        temp_buffer.seek(0)

        wb = openpyxl.load_workbook(temp_buffer)
        ws = wb.active

        # DATE FORMAT APPLY
        for col_name in date_columns:
            if col_name in final_df.columns:
                col_index = list(final_df.columns).index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_index)
                    if cell.value:
                        cell.number_format = "DD-MM-YYYY"

        # NUMBER FORMAT FIX
        for col in range(1, ws.max_column + 1):
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0"

        # =============================
        # 1️⃣4️⃣ FINAL SAVE → SEND
        # =============================
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        return send_file(
            output_buffer,
            as_attachment=True,
            download_name="Final_Output.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return render_template_string(HTML_FORM,
                                      error=f"Error: {str(e)}",
                                      success=None)


# =============================
# VERCEL ENTRY POINT
# =============================
# Vercel uses the `app` object directly (WSGI).
# For local testing: python app.py
if __name__ == "__main__":
    app.run(debug=True)
